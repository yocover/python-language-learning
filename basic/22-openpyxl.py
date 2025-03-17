# 使用python 中的第三方库 openpyxl
# openpyxl 是一个用于读写 Excel 文件的库
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import random

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# 写入excel 文件

# 创建工作簿
wb = openpyxl.Workbook()

# 创建工作表
sheet = wb.active
sheet.title = "期末成绩"

titles = (
    "姓名",
    "数学",
    "语文",
    "英语",
    "物理",
    "化学",
    "生物",
    "政治",
    "历史",
    "地理",
)

for col_index, title in enumerate(titles):
    sheet.cell(row=1, column=col_index + 1, value=title)


names = ("张三", "李四", "王五", "赵六", "孙七", "周八", "吴九", "郑十")

for row_index, name in enumerate(names):
    sheet.cell(row=row_index + 2, column=1, value=name)
    for col_index in range(2, len(titles) + 1):
        sheet.cell(row=row_index + 2, column=col_index, value=random.randint(0, 100))


# 调整样式和共识计算


# 保存
wb.save("期末成绩.xlsx")


# 对齐方式
alignment = Alignment(horizontal="center", vertical="center")
# 边框线条
side = Side(color="ff7f50", style="mediumDashed")

wb = openpyxl.load_workbook("期末成绩.xlsx")
sheet = wb.worksheets[0]

# 调整行高和列宽
sheet.row_dimensions[1].height = 30
sheet.column_dimensions["E"].width = 120

sheet["E1"] = "平均分"
# 设置字体
sheet.cell(1, 5).font = Font(size=18, bold=True, color="ff1493", name="华文楷体")
# 设置对齐方式
sheet.cell(1, 5).alignment = alignment
# 设置单元格边框
sheet.cell(1, 5).border = Border(left=side, top=side, right=side, bottom=side)
for i in range(2, len(names) + 2):
    # 公式计算每个学生的平均分
    sheet[f"E{i}"] = f"=average(B{i}:D{i})"
    sheet.cell(i, 5).font = Font(size=12, color="4169e1", italic=True)
    sheet.cell(i, 5).alignment = alignment

wb.save("期末成绩.xlsx")

# 生成统计图表

wb = Workbook(write_only=True)
sheet = wb.create_sheet()

rows = [
    ("类别", "销售A组", "销售B组"),
    ("手机", 40, 30),
    ("平板", 50, 60),
    ("笔记本", 80, 70),
    ("外围设备", 20, 10),
]

# 向表单中添加行
for row in rows:
    sheet.append(row)

# 创建图表对象
chart = BarChart()
chart.type = "col"
chart.style = 10
# 设置图表的标题
chart.title = "销售统计图"
# 设置图表纵轴的标题
chart.y_axis.title = "销量"
# 设置图表横轴的标题
chart.x_axis.title = "商品类别"
# 设置数据的范围
data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=3)
# 设置分类的范围
cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
# 给图表添加数据
chart.add_data(data, titles_from_data=True)
# 给图表设置分类
chart.set_categories(cats)
chart.shape = 4
# 将图表添加到表单指定的单元格中
sheet.add_chart(chart, "A10")

wb.save("demo.xlsx")
